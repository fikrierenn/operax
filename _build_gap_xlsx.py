# -*- coding: utf-8 -*-
"""Mikro ↔ Operax kapsamlı GAP analizi Excel üretici (tek seferlik)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Calibri"
wb = Workbook()

# --- stiller ---
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="1F4E78")        # koyu mavi başlık
SUB = Font(name=FONT, bold=True, size=12, color="1F4E78")
NORM = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# lazım renkleri
FILL_EVET = PatternFill("solid", fgColor="C6EFCE")   # yeşil
FILL_OLUR = PatternFill("solid", fgColor="FFEB9C")   # sarı
FILL_HAYIR = PatternFill("solid", fgColor="F2F2F2")  # gri
FILL_VAR = PatternFill("solid", fgColor="D9E1F2")    # açık mavi (zaten var)
def lazim_fill(v):
    v=(v or "").upper()
    if v.startswith("VAR"): return FILL_VAR
    if "EVET" in v: return FILL_EVET
    if "OLUR" in v: return FILL_OLUR
    if "HAYIR" in v: return FILL_HAYIR
    return None

def sheet(title, headers, rows, widths, freeze="A2", lazim_col=None, note=None):
    ws = wb.create_sheet(title)
    r0 = 1
    if note:
        ws.cell(1,1,note).font = Font(name=FONT, italic=True, size=9, color="808080")
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
        ws.cell(1,1).alignment = WRAP
        r0 = 2
    for c,h in enumerate(headers,1):
        cell = ws.cell(r0,c,h); cell.font=H; cell.fill=HF; cell.alignment=CTR; cell.border=BORDER
    for ri,row in enumerate(rows, r0+1):
        for c,val in enumerate(row,1):
            cell = ws.cell(ri,c,val); cell.font=NORM; cell.alignment=TOP; cell.border=BORDER
            if lazim_col and c==lazim_col:
                f=lazim_fill(val)
                if f: cell.fill=f
                cell.alignment=CTR; cell.font=Font(name=FONT,bold=True,size=10)
    for c,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws[freeze] if not note else ws["A3"]
    ws.row_dimensions[r0].height = 28
    return ws

# ============ 0. KAPAK / OKUMA KILAVUZU ============
ws = wb.active; ws.title = "00_Kapak"
ws["A1"]="Operax ↔ Mikro / ERP — Kapsamlı GAP Analizi (Evrak · Hareket · Kalem · Defter)"
ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1F4E78")
ws.merge_cells("A1:F1")
meta = [
 ("Tarih","2026-05-30"),
 ("Kaynaklar","Mikro V17 (ozgurguler.net mirror) + Fikri resmi doküman yapıştırma + ERPNext/Smartstore/nopCommerce/RealAhmed WMS"),
 ("Operax envanteri","Lib/Dtos.cs (code-explorer, kesin) + docs/sql/*"),
 ("İlgili belgeler","docs/MIKRO_V16_ANALYSIS.md (§0.5,§12,§13) · docs/REFERENCE_STUDY.md (B1-B18) · plans/12-16"),
 ("Kanıt notu","[REPO-HTM]=Mikro sayfası okundu · [OPERAX]=Dtos.cs · DOĞRULANMADI=teyit edilmemiş, tahmin değil"),
]
r=3
for k,v in meta:
    ws.cell(r,1,k).font=Font(name=FONT,bold=True,size=10)
    ws.cell(r,2,v).font=NORM; ws.cell(r,2).alignment=WRAP
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    r+=1
r+=1
ws.cell(r,1,"LAZIM? RENK KODU").font=SUB; r+=1
for txt,fill in [("VAR — Operax'ta mevcut",FILL_VAR),("EVET — üretilmeli (gerçek ihtiyaç)",FILL_EVET),
                 ("OLUR — sektöre/ileriye bağlı",FILL_OLUR),("HAYIR-ŞİMDİ — ertelenmiş/kapsam dışı",FILL_HAYIR)]:
    c=ws.cell(r,1,txt); c.fill=fill; c.font=NORM; c.border=BORDER
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
r+=1
ws.cell(r,1,"SHEET REHBERİ").font=SUB; r+=1
for s,d in [
 ("01_Stok_Hareket_Tipi","Mikro sth_cins/sth_evraktip ↔ Operax MovementType/SourceDoc"),
 ("02_Cari_Hareket_Tipi","Mikro cha_cinsi/cha_evrak_tip ↔ Operax AccountMovementType"),
 ("03_Cek_Senet_Durum","Mikro sck_sonpoz/sck_tip ↔ Operax Cheque statü"),
 ("04_Belge_Modulleri","Mikro belge/sipariş tipleri ↔ Operax Features ekranları"),
 ("05_Kalem_Hizmet_Masraf","Mal/Hizmet/Masraf/Dönemsel gider saklama (Mikro §13)"),
 ("06_Muhasebe_GL","Hesap planı/posting-rule/masraf merkezi (Mikro §3.5, K1)"),
 ("07_ERP_Karma_Dersler","ERPNext/Smartstore/nop'tan alınan + karma öneriler"),
 ("08_Backlog_Onceli","Tüm GAP'ler tek tablo — öncelik/etki/maliyet/plan bağı"),
]:
    ws.cell(r,1,s).font=Font(name=FONT,bold=True,size=10,color="1F4E78")
    ws.cell(r,2,d).font=NORM; ws.cell(r,2).alignment=WRAP
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); r+=1
for c,w in zip("ABCDEF",[26,30,20,18,18,18]): ws.column_dimensions[c].width=w
ws.column_dimensions["B"].width=46

# ============ 01 STOK HAREKET TİPİ ============
sheet("01_Stok_Hareket_Tipi",
 ["Mikro sth_cins / evraktip","Açıklama","Operax karşılığı","Lazım?","Operax çözüm / not","Plan/Backlog"],
 [
  ["0:Toptan / 1:Perakende","Toptan/perakende mal giriş-çıkış","RECEIPT/ISSUE + SourceDoc","VAR","Perakende-POS ayrı değil (E3)","-"],
  ["sth_evraktip 13:Giriş İrsaliyesi","Mal kabul irsaliyesi (mali belge değil)","Receiving ≈ irsaliye","EVET","İrsaliye↔Fatura AYRIMI YOK; VUK: mal hareketi=irsaliye, mali=fatura. İrsaliyeden faturaya dönüşüm zinciri kur","B17/E1"],
  ["sth_evraktip 1:Çıkış İrsaliyesi","Sevk irsaliyesi","Shipping ≈ irsaliye","EVET","Aynı E1; sevk irsaliyesi → satış faturası dönüşümü","B17/E1"],
  ["sth_evraktip 3:Giriş Faturası / 4:Çıkış Faturası","Alış/Satış faturası (mali)","EI / SI","VAR","İrsaliye-fatura zinciri kurulmalı (E1)","B17/E1"],
  ["normal_iade=1 (her tip)","Alış/Satış iade","YOK (ADJUST'a karışır)","EVET","İade ayrı belge: orijinale bağ + ters-kayıt (immutability). SourceDocType=RETURN_IN/OUT","B17/E2"],
  ["4:Fire","Fire/zayi/imha","YOK","EVET","ADJUST + AdjustReason=WASTE/SCRAP; maliyet+vergi etkisi ayrı","B17/E4"],
  ["5:Sarf","Üretim-dışı sarf/gider sarfı","ISSUE (sadece üretim)","EVET","SourceDocType=CONSUMPTION; üretim dışı sarf evrakı","B17/E9"],
  ["10:Sayım","Sayım fark düzeltme","COUNT_ADJ","VAR","Fazla/eksik AYRIMI yok → COUNT_PLUS/COUNT_MINUS sebep","B17/E5"],
  ["11:Stok Açılış","Dönem başı / go-live stok yükleme","YOK","EVET","SourceDocType=OPENING_STOCK (AccountMovement'ta OPENING var, stokta yok)","B17/E7"],
  ["3:Stok Virman","Aynı depo birim/lot/raf düzeltme","YOK (TRANSFER depo-arası)","OLUR","Depo-içi virman; düşük öncelik","-"],
  ["6:Transfer / evraktip 2","Depo-arası transfer","TRANSFER","VAR","-","-"],
  ["7:Üretim / evraktip 7","Üretimden mamul giriş","PRODUCTION","VAR","Sarf (ISSUE) + mamul (PRODUCTION) ayrı kimlik zayıf","B17/E9"],
  ["8:Fason","Fason giriş/çıkış","YOK","OLUR","Fason üretim yapılırsa","-"],
  ["2:Dış Ticaret / 12:İthalat-İhracat","İthalat/ihracat + antrepo","YOK","OLUR","GTİP/gümrük + landed cost (ithalat masraf dağıtımı)","-"],
  ["9:Değer Farkı","Stok değerleme/maliyet düzeltme","VARIANCE (cari) — stokta yok","OLUR","Enflasyon/maliyet düzeltme hareketi","-"],
  ["evraktip 15:Depolar Arası Satış","Şubeler arası satış","YOK","OLUR","Intercompany (VISION §7.5 parametrik)","-"],
  ["13:Hal / 14:Müstahsil(?)","Hal/tarım niş","YOK","HAYIR-ŞİMDİ","Niş sektör; 14/15 DOĞRULANMADI","-"],
  ["evraktip 18:Demirbaşa Virman","Demirbaş hareketi","YOK","HAYIR-ŞİMDİ","Demirbaş modülü yok","-"],
 ],
 [34,34,26,12,46,14], lazim_col=4,
 note="Operax MEVCUT MovementType (5): RECEIPT·ISSUE·TRANSFER·COUNT_ADJ·PRODUCTION | SourceDoc (6): RECEIVING·SHIPPING·TRANSFER·COUNT·PRODUCTION·PICKING  [OPERAX Dtos.cs]. Mikro sth_cins 0-13 [REPO-HTM]; 14/15 DOĞRULANMADI.")

# ============ 02 CARİ HAREKET TİPİ ============
sheet("02_Cari_Hareket_Tipi",
 ["Mikro cha_cinsi / evrak_tip","Açıklama","Operax karşılığı","Lazım?","Operax çözüm / not","Plan/Backlog"],
 [
  ["6:Toptan / 7:Perakende / 8:Hizmet Faturası","Satış/alış/hizmet faturası","SALES_INVOICE / PURCHASE_INVOICE","VAR","Hizmet faturası: Item.ItemKind=SERVICE (B18)","B18"],
  ["evrak_tip 1:Tahsilat Makbuzu","Müşteriden tahsilat","COLLECTION","VAR","-","-"],
  ["evrak_tip 37:Kasa Masraf Fişi / 55:Giriş Gider Makbuzu","Gider/masraf belgesi","YOK (Expenses basic)","EVET","Gider fişi belgesi + AccountMovement besleme","B18/plan16"],
  ["Tediye/Ödeme (kasa/banka)","Tedarikçiye ödeme","PAYMENT","VAR","-","-"],
  ["1:Müşteri Çeki / 4:Çek Giriş Bordrosu","Çek giriş","CHEQUE_IN + Cheque","VAR","-","-"],
  ["Verilen Çek / Çek Çıkış Bordrosu","Çek çıkış/ciro","CHEQUE_OUT","VAR","-","-"],
  ["2:Müşteri Senedi / Senet Bordroları","Senet giriş/çıkış","PromissoryNote","VAR kısmi","Senet modülü var","-"],
  ["16:Cari Açılışı / evrak_tip 29:Açılış Fişi","Cari devir bakiye","OPENING","VAR","-","-"],
  ["10:Vade Farkı Faturası","Geç ödeme vade farkı","YOK","EVET","TR yaygın; AccountMovementType=LATE_FEE","B17/E12"],
  ["evrak_tip 33:Genel Virman Dekontu","Cari↔cari / hesap↔hesap virman","YOK (TRANSFER_IN/OUT sabit var, evrak yok)","EVET","Virman evrakı — Plan 11 başlamadı","B17/E11/plan11"],
  ["31:Borç Dekontu / 32:Alacak Dekontu","Serbest borç/alacak düzeltme","YOK (sadece VARIANCE)","EVET","Dekont belgesi + AccountMovementType","B17/E12"],
  ["11:Kur Farkı Faturası","Dövizli cari kur değerleme","YOK","OLUR","Dövizli çalışınca (E10); çok-döviz kolonu da gerekli","B17/E10"],
  ["34:Gelen Havale / 35:Gönderilen Havale","Banka havale/EFT","TransactionType TRANSFER var, evrak yok","OLUR","Banka havale evrakı","-"],
  ["33:Avans Makbuzu","Müşteri/tedarikçi avansı","YOK","OLUR","Avans takibi","-"],
  ["9:Serbest Meslek Makbuzu","SMM (danışman/avukat)","YOK","OLUR","Tevkifatlı SMM","-"],
  ["25:Müstahsil / 24:Hal / 29:Gümrük Beyan","Sektör/mevzuat belgesi","YOK","OLUR","Sektör bağlı","-"],
  ["37:Teminat Mektubu / 39:Depozito","İleri finans araçları","YOK","HAYIR-ŞİMDİ","İleri finans","-"],
 ],
 [38,30,30,12,40,16], lazim_col=4,
 note="Operax MEVCUT AccountMovementType (9): SALES_INVOICE·PURCHASE_INVOICE·PAYMENT·COLLECTION·CHEQUE_IN·CHEQUE_OUT·OPENING·VARIANCE·REVERSAL [OPERAX]. Mikro cha_cinsi 0-41 [REPO-HTM]; cha_evrak_tip 0-50 teyitli, 51-137 DOĞRULANMADI.")

# ============ 03 ÇEK/SENET DURUM ============
sheet("03_Cek_Senet_Durum",
 ["Mikro sck_sonpoz / sck_tip","Açıklama","Operax Cheque statü","Lazım?","Not"],
 [
  ["sonpoz 0:Portföyde","Elde, işlem görmemiş","PORTFOLIO","VAR","-"],
  ["sonpoz 1:Ciro","Başkasına ciro","ENDORSED","VAR","-"],
  ["sonpoz 2:Tahsilde","Bankaya tahsile verildi","IN_BANK","VAR","-"],
  ["sonpoz 10:Ödendi","Tahsil/ödeme tamam","COLLECTED / PAID","VAR","-"],
  ["sonpoz 4:İade / 7:Ödenmedi İade","Karşılıksız iade","RETURNED","VAR kısmi","-"],
  ["sonpoz 3:Teminatta","Çek/senet teminata verme","YOK","EVET","TR yaygın; yeni statü TEMINATTA → document-immutability §2.4"],
  ["sonpoz 9:Kısmen Ödendi","Kısmi tahsilat","YOK","EVET","sck_odenen kolonu; PARTIALLY_PAID statü + kalan tutar"],
  ["sonpoz 8:İcrada","Karşılıksız sonrası icra","YOK","OLUR","İcra takibi"],
  ["sonpoz 6:Ödenmedi Portföyde","Vade geçti hâlâ elde","YOK","OLUR","Vadesi geçmiş uyarı"],
  ["sck_tip 0-13 (müşteri/kendi çek-senet, teminat, depozito)","14 araç tipi","Direction RECEIVED/ISSUED","VAR kısmi","Teminat mektubu/depozito tipleri yok"],
  ["sck_nerede_cari_kodu","Çek şu an kimde/hangi banka","YOK","EVET","Çek KONUM izleme — Operax GAP (MIKRO §6)"],
 ],
 [40,30,24,12,46], lazim_col=4,
 note="Operax MEVCUT Cheque statü (6): PORTFOLIO·IN_BANK·COLLECTED·RETURNED·ENDORSED·PAID [OPERAX]. Mikro sck_sonpoz 0-10, sck_tip 0-13 [REPO-HTM].")

# ============ 04 BELGE MODÜLLERİ ============
sheet("04_Belge_Modulleri",
 ["Belge/Modül","Mikro","Operax Features","Lazım?","Not"],
 [
  ["Satınalma Sipariş","SIPARISLER sip_cins=7 (Satınalma Talebi)","PurchaseOrders","VAR","-"],
  ["Satış Sipariş","SIPARISLER (Normal)","SalesOrders","VAR","-"],
  ["Mal Kabul / Giriş İrsaliye","sth_evraktip 13","Receiving","VAR","İrsaliye-fatura ayrımı E1"],
  ["Sevkiyat / Çıkış İrsaliye","sth_evraktip 1","Shipping","VAR","E1"],
  ["Alış Faturası","sth_evraktip 3","SalesInvoices(EI)","VAR","-"],
  ["Satış Faturası","sth_evraktip 4","SalesInvoices(SI)","VAR","-"],
  ["Depo Transfer","sth_evraktip 2","Transfer","VAR","-"],
  ["Sayım","SAYIM_SONUCLARI","CycleCount","VAR","Freeze K5 (M08 spec)"],
  ["Üretim/İş Emri","ISEMIRLERI + sip_cins 9","Production/Manufacturing","VAR","-"],
  ["Picking","-","Picking","VAR","Operax fazlası (WMS)"],
  ["Lot/Serial/LPN izleme","PARTILOT + STOK_SERINO","Lot/Serial/LPN","VAR","Operax fazlası (WMS)"],
  ["Kasa/Banka/Çek/Kredi/Kart","KASALAR/BANKALAR/ODEME_EMIRLERI","Finance/*","VAR","-"],
  ["İade Belgesi","normal_iade flag","YOK","EVET","Ayrı iade belgesi (E2)"],
  ["Virman Evrakı","Genel Virman Dekontu","YOK","EVET","Plan 11 (E11)"],
  ["Konsinye","sip_cins=1","YOK","OLUR","Konsinye giriş/çıkış (E6)"],
  ["Proforma","PROFORMA_SIPARISLER","YOK","OLUR","Proforma teklif"],
  ["Teklif (RFQ)","VERILEN/ALINAN_TEKLIFLER","YOK","OLUR","M03 spec'te var (RFQ)"],
  ["B2B Portal","-","M11 Portal (TODO)","OLUR","Operax roadmap"],
  ["Servis/RMA","BAKIM/ARIZA tabloları","Service (boş)","OLUR","M12"],
  ["Demirbaş/Amortisman","DEMIRBASLAR","YOK","HAYIR-ŞİMDİ","Sabit kıymet modülü"],
  ["Personel/Bordro","PERSONELLER/TAHAKKUK","YOK","HAYIR-ŞİMDİ","İK/bordro kapsam dışı"],
  ["Resmi Muhasebe/GL","MUHASEBE_FISLERI","YOK (ertelendi K1/K2)","HAYIR-ŞİMDİ","Periyodik GL modülü, mevzuat skill ön koşul"],
 ],
 [28,32,26,12,34], lazim_col=4,
 note="Operax 28 Features klasörü [OPERAX code-explorer]. Operax WMS tarafı (Picking/Lot/Serial/LPN) Mikro'dan ZENGİN; Mikro muhasebe/demirbaş/personel tarafı zengin.")

# ============ 05 KALEM / HİZMET / MASRAF ============
sheet("05_Kalem_Hizmet_Masraf",
 ["Kavram","Mikro","Operax","Lazım?","Operax çözüm / not"],
 [
  ["Mal/Ürün","STOKLAR (stok bakiyeli)","Item (ItemType base şemada yok)","VAR","ItemKind=GOODS"],
  ["Hizmet","HIZMET_HESAPLARI ayrı kart (satış+alış GL, tevkifat, stoksuz)","YOK (ItemType='SERVICE' planlı, SP kullanmıyor)","EVET","Item.ItemKind=SERVICE; onay SP guard → StockMovement YAZMA"],
  ["Masraf","MASRAF_HESAPLARI ayrı kart (gider GL+KKEG, satış kodu yok)","ExpenseType zayıf (GL kodu yok)","EVET","ExpenseType'a AccountCode+Direction+KDV+IsKkeg ekle"],
  ["Masraf Merkezi","SORUMLULUK_MERKEZLERI (cost center, dağıtım)","CostCenter var (tek boyut, dağıtım yok)","OLUR","Plan 06 çok-boyut+%dağıtım; K1 GL ile"],
  ["Dönemsel Gider","DONEMLERE_YAYILAN_HIZMETLER (peşin kira/sigorta, başlangıç-bitiş+GL1/GL2)","YOK","OLUR","DeferredExpense tablosu + aylık tahakkuk Hangfire job"],
  ["Hizmet/Gider hareketi","CARI_HESAP_HAREKETLERI (cha_cinsi 8/9, evrak_tip 37/55) — stok değil","ExpenseInvoice (AccountMovement'a beslenmiyor)","EVET","Gider fişi → AccountMovement atomik (plan 16)"],
  ["Tevkifat / KKEG / ÖİV","Hizmet+masraf kartında kolon","YOK","OLUR","VUK gider sınıflaması; mevzuat skill ile (K2)"],
  ["Stok kalemi GL eşleme","STOK_MUHASEBE_GRUPLARI (grup→~23 GL kodu)","YOK","HAYIR-ŞİMDİ","K1 posting-rule (PostingRule tablosu)"],
 ],
 [22,46,34,12,40], lazim_col=4,
 note="Mikro: mal/hizmet/masraf 3 AYRI master (stoksuz kalem stok defterini kirletmez). Operax önerisi: tek Item + ItemKind discriminator (GOODS/SERVICE/EXPENSE) + onay SP'de stoksuz-kalem guard. Kolon detayı DOĞRULANMADI (mirror ECONNREFUSED).")

# ============ 06 MUHASEBE / GL ============
sheet("06_Muhasebe_GL",
 ["Kavram","Mikro","Operax","Lazım?","Not"],
 [
  ["Hesap Planı (kebir)","MUHASEBE_HESAP_PLANI (hesap_tip Aktif/Pasif/Gelir/Gider/Nazım + çalışma şekli + hiyerarşi)","YOK","HAYIR-ŞİMDİ","K1/K2 GL modülü; tekdüzen hesap planı"],
  ["Yevmiye Fişi","MUHASEBE_FISLERI (fis_meblag0 işaretli ±, yevmiye no, fis_tur 0-4)","YOK","HAYIR-ŞİMDİ","K1; periyodik posting (gerçek-zamanlı değil)"],
  ["Subledger→GL köprü","fis_ticari_uid + fis_ticari_tip (gevşek bağ, perpetual DEĞİL)","AccountMovement subledger var, GL yok","HAYIR-ŞİMDİ","K1/K3 — Operax cari subledger DOĞRU yolda"],
  ["Posting-Rule eşleme","STOK_MUHASEBE_GRUPLARI (grup+hareket tipi→GL hesap kodu)","YOK","HAYIR-ŞİMDİ","K1 NET DESEN: PostingRule(grup,hareket,hesap) normalize"],
  ["Masraf merkezi/proje boyutu","SORUMLULUK_MERKEZLERI + fis_proje_kodu","CostCenter var (kısmi)","OLUR","Plan 06 + K1"],
  ["Mahsup türleri","fis_fmahsup_tipi 23 değer (SMM/kur farkı/enflasyon/dönem kapanış…)","YOK","HAYIR-ŞİMDİ","K1 otomatik mahsup kataloğu"],
  ["Dönem kilidi","muh_kilittarihi (hesap-bazlı) + firma+mali yıl","YOK (plan 14 K4 zaman-bazlı geliyor)","EVET","K4 dönem kontrolü (plan 14); hesap-bazlı kilit Operax'ta yok"],
  ["Çok-döviz (yerli/alt/orj)","fis_meblag0/1/2 paralel","Tek para (TRY)","OLUR","Dövizli çalışınca"],
  ["e-Defter / Berat / GİB","DEFTER_BEYAN tabloları","YOK","HAYIR-ŞİMDİ","K5: Operax ÜRETMEZ, sadece LOCKED döneme saygı"],
 ],
 [24,46,30,12,40], lazim_col=4,
 note="Mikro muhasebe tarafı = Operax'ın ERTELENMİŞ K1/K2 modülü için BİREBİR referans. Operax kararı: subledger gerçek-zamanlı + GL periyodik posting (gerçek-zamanlı GL DEĞİL). Mevzuat skill (K2) ön koşul. e-Defter üretimi kapsam dışı (K5).")

# ============ 07 ERP KARMA DERSLER ============
sheet("07_ERP_Karma_Dersler",
 ["Kaynak ERP","Desen / Ders","Operax'a karma öneri","Lazım?","Backlog"],
 [
  ["ERPNext","Stock Ledger Entry + GL Entry ikilisi (perpetual)","Operax subledger ayrı tut AMA periyodik posting (gerçek-zamanlı GL değil) — Mikro+ERPNext karması","HAYIR-ŞİMDİ","K1"],
  ["ERPNext","stock_queue JSON (FIFO kuyruk)","Mikro kalıcı eşleme tablosu TERCİH edildi (StockCostConsumption) — denetlenebilir","EVET","B5/K7"],
  ["ERPNext","Immutable ledger + is_cancelled (reversal, silme yok)","Operax IsCancelled + reversal (plan 14); Mikro _iptal de aynı yönde","EVET","plan14"],
  ["ERPNext","against_voucher (ödeme→fatura mahsup)","Açık-kalem kapama tablosu (Mikro CARI_HAREKET_BORC_ALACAK_ESLEME ile birleşik ders)","OLUR","B16"],
  ["Smartstore / nopCommerce","Multi-store izolasyon — EF global filter YOK, elle","Operax CompanyId her satır + TVF @CompanyId-sargı + analyzer guard (Mikro firmano deseni de aynı)","EVET","B1/plan12"],
  ["Smartstore","Modüler/plugin sınırları","Operax Feature-based modül + Module aktivasyon (mevcut)","VAR","-"],
  ["RealAhmed WMS (.NET)","Available vs Allocated stok ayrımı","tvf_InventoryBalance rezervasyon kolonu","OLUR","B6"],
  ["RealAhmed WMS","Lokasyon IsReceivable/IsPickable eligibility","Hedef hücre guard","OLUR","B8"],
  ["ModernWMS","ASN (sevk öncesi bildirim)","3PL/tedarikçi entegrasyon senaryosu","HAYIR-ŞİMDİ","B10"],
  ["Slice (Transaction Script)","Vertical slice + Dapper + command-query","Operax zaten aynı felsefe — slice tutarlılık audit","OLUR","B9"],
  ["Mikro","Polymorphic ledger (tek hareket tablosu + tip)","Operax StockMovement/AccountMovement+tip DOĞRULANDI; yeni belge=yeni tip, yeni tablo DEĞİL","VAR","§0.5"],
  ["Mikro","Belge başlığı ledger'a gömme","REDDEDİLDİ — Operax ayrı Header/Line (normalize, immutability+durum makinesi)","VAR","§0.5"],
  ["Mikro","Çek konum izleme (nerede_cari)","Operax çek modülüne 'şu an kimde' alanı","EVET","§6"],
  ["KARMA SONUÇ","WMS: Operax > Mikro (Picking/Lot/Serial/Wave). Muhasebe: Mikro > Operax (GL/hesap planı). Cari: benzer.","Operax = güçlü WMS+operasyon + hafif cari (K3) + ertelenmiş periyodik GL (K1). Mikro'dan muhasebe deseni, ERPNext'ten ledger disiplini, kendi WMS üstünlüğü.","-","-"],
 ],
 [20,42,46,12,12], lazim_col=4,
 note="Karma strateji: her ERP'nin güçlü yanı. WMS=Operax/RealAhmed, Ledger immutability=ERPNext, Muhasebe/posting=Mikro, İzolasyon=hepsi elle (Operax CompanyId her satır).")

# ============ 08 BACKLOG ÖNCELİK ============
sheet("08_Backlog_Onceli",
 ["Kod","İş","Kaynak","Lazım?","Etki","Maliyet","Plan/Durum"],
 [
  ["E1","İrsaliye↔Fatura ayrımı + dönüşüm zinciri","Mikro sth_evraktip","EVET","Yüksek","Orta","B17 — yeni"],
  ["E2","Alış/Satış İade belgesi (ters-kayıt)","Mikro normal_iade","EVET","Yüksek","Orta","B17 — yeni"],
  ["E4","Fire/Zayi/İmha (ADJUST sebep kodu)","Mikro sth_cins 4","EVET","Yüksek","Düşük","B17 — yeni"],
  ["E11","Virman (kasa↔kasa, cari↔cari)","Mikro Genel Virman","EVET","Yüksek","Orta","Plan 11 (başlamadı)"],
  ["B1","Multi-company izolasyon (TVF-sargı+analyzer)","tüm ERP + Mikro firmano","EVET","Existential","Düşük","Plan 12 (onaylı)"],
  ["plan14","Immutability + dönem kontrolü K4 + K8 istisna/iz + clustered PK","ERPNext+Mikro+VUK","EVET","Yüksek","Orta","Plan 14 (onay bekliyor)"],
  ["B3","Hafif cari besleme (AccountMovement)","R0 + Mikro subledger","EVET","Yüksek","Orta","Plan 16 (taslak)"],
  ["B18","Hizmet/Masraf kalem tipi (Item.ItemKind)","Mikro HIZMET/MASRAF_HESAPLARI","EVET","Yüksek","Orta","yeni; Plan 06 ile"],
  ["B5","FIFO kalıcı StockCostConsumption eşleme","ERPNext+Mikro §1.5","EVET","Orta","Orta","K7 (revize)"],
  ["E5","Sayım fazla/eksik ayrımı","Mikro sth_cins 10","EVET","Orta","Düşük","B17"],
  ["E7","Stok açılış/devir fişi","Mikro sth_cins 11","EVET","Orta","Düşük","B17"],
  ["E12","Vade farkı + borç/alacak dekontu","Mikro cha 10/dekont","EVET","Orta","Orta","B17"],
  ["Çek+","Çek statü: TEMİNAT + KISMİ ÖDEME + konum izleme","Mikro sck_sonpoz 3/9 + nerede_cari","EVET","Orta","Düşük","B17/§6"],
  ["B16","Açık-kalem kapama tablosu (AccountReconciliation)","Mikro CARI_HAREKET_BORC_ALACAK_ESLEME","OLUR","Orta","Orta","yeni"],
  ["B6","Available vs Allocated stok","RealAhmed WMS","OLUR","Orta","Düşük","backlog"],
  ["E3/E6/E8/E10","Perakende/Konsinye/Fason/Kur farkı","Mikro cins","OLUR","Düşük-Orta","Orta","sektöre bağlı"],
  ["Defer","Dönemsel gider (DeferredExpense + job)","Mikro DONEMLERE_YAYILAN","OLUR","Düşük","Orta","ertele"],
  ["K1/K2","Periyodik GL muhasebeleştirme modülü","Mikro muhasebe + mevzuat skill","HAYIR-ŞİMDİ","Yüksek","Yüksek","ertelendi; mevzuat skill ön koşul"],
  ["E13","GL Mahsup/Açılış/Kapanış","Mikro fis_tur/fmahsup","HAYIR-ŞİMDİ","Yüksek","Yüksek","K1 ile"],
  ["Demirbaş/Personel","Sabit kıymet + bordro","Mikro","HAYIR-ŞİMDİ","Düşük","Yüksek","kapsam dışı"],
 ],
 [10,42,30,11,12,11,22], lazim_col=4,
 note="ÖNCELİK SIRASI (üstten): EVET-Yüksek üretilmeli → OLUR sektör/ileri → HAYIR-ŞİMDİ ertelenmiş. En yüksek 4: E1,E2,E4,E11. Mevcut onaylı: plan 12/14/16.")

import os
out = r"D:/Dev/Operax/docs/Operax_Mikro_GAP_Analizi.xlsx"
wb.save(out)
print("KAYDEDILDI:", out, "| sheet:", wb.sheetnames)
